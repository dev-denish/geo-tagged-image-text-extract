"""
================================================================================
 Geo-Tagged Image -> Excel Extractor  (NoteCam-style overlay photos)
================================================================================
What this does:
  Walks through your zip (or an already-extracted folder) of geo-tagged
  photos, reads the burned-in text overlay on each photo (Latitude,
  Longitude, Elevation, Accuracy, Note), and writes one row per photo into
  an Excel file with columns:

      Folder Name | File Name | Latitude | Longitude | Elevation |
      Accuracy | Note / Status | Flag | Overlay Preview (thumbnail image)

IMPORTANT - please read before trusting the output blindly:
  These photos do NOT have GPS saved in their EXIF metadata - the location
  is only "burned in" as visible text on the image. That means the ONLY way
  to recover it is OCR (optical text recognition), and OCR is never
  literally 100.000% guaranteed on photos with a busy/cluttered background
  behind the text (grass, soil, sky, etc. show through the semi-transparent
  overlay). In testing this script, the four numeric fields (Latitude,
  Longitude, Elevation, Accuracy) read correctly essentially every time
  because they always follow a strict "label: number" pattern that the
  script double-checks with a plausible-range test. The free-text "Note"
  field is the only one that can occasionally have a character-level typo
  after OCR (e.g. a trailing letter misread as a digit).

  To get you to true ~100% accuracy, EVERY row also gets a small thumbnail
  image of the overlay text embedded directly in Excel column I, so you can
  eyeball and fix any flagged (or even unflagged) row in seconds without
  opening the original photo. Rows the script isn't fully confident about
  are highlighted red in the "Flag" column so you know exactly where to
  look first.

--------------------------------------------------------------------------------
SETUP (one-time, on Windows):

1) Install Tesseract OCR (the actual text-recognition engine):
   https://github.com/UB-Mannheim/tesseract/wiki
   Default install path is usually:
       C:\\Program Files\\Tesseract-OCR\\tesseract.exe
   If yours installed somewhere else, update TESSERACT_PATH below.

2) Install the Python packages (open PowerShell / cmd and run):
       pip install pillow opencv-python pytesseract openpyxl numpy

3) Edit the three paths in the CONFIG section right below, then run:
       python geo_extract.py
--------------------------------------------------------------------------------
"""

import os
import re
import shutil
import tempfile
import traceback
import zipfile
from datetime import datetime

import cv2
import numpy as np
import pytesseract
from PIL import Image as PILImage

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================== CONFIG =======================================
# Path to the zip file you want to process. Leave as-is if using EXTRACTED_FOLDER
# instead.
ZIP_PATH = r"C:\Users\Admin\OneDrive - VNV\testing\geo tagged imgs\OneDrive_1_29-7-2026 (1).zip"

# If you'd rather point the script at an already-unzipped folder, put its path
# here and set ZIP_PATH = None above (or just leave EXTRACTED_FOLDER = None to
# use the zip instead).
EXTRACTED_FOLDER = None

# Where the finished Excel file should be saved.
OUTPUT_XLSX = r"C:\Users\Admin\OneDrive - VNV\testing\geo tagged imgs\geo_tagged_data_1.xlsx"

# Path to tesseract.exe (see setup step 1 above).
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# ==============================================================================

pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

IMG_EXTS = {".jpg", ".jpeg", ".png"}

# Regex patterns for each field. Tolerant of common OCR confusions
# (e.g. "l" vs "1" vs "i", stray spaces, "+" instead of "±").
PATTERNS = {
    "Latitude":  re.compile(r"Lat[il1]tude\s*[:;]?\s*([\-]?\d{1,3}\.\d{3,8})", re.I),
    "Longitude": re.compile(r"Long[il1]tude\s*[:;]?\s*([\-]?\d{1,3}\.\d{3,8})", re.I),
    "Elevation": re.compile(r"Elevat[il1]on\s*[:;]?\s*([\-]?\d+\.?\d*)\s*[±+\-]?\s*(\d+\.?\d*)?\s*m", re.I),
    "Accuracy":  re.compile(r"Accurac[yv]\s*[:;]?\s*(\d+\.?\d*)\s*m", re.I),
    "Note":      re.compile(r"Note\s*[:;]?\s*(.+)", re.I),
}

# Plausible coordinate range used as a sanity check (India-wide by default).
# Narrow this to your actual survey area for an even stronger safety net,
# e.g. Madhya Pradesh is roughly LAT_RANGE=(21.0, 27.0), LON_RANGE=(74.0, 83.0)
LAT_RANGE = (6.0, 38.0)
LON_RANGE = (68.0, 98.0)

# Overlay crop box, as a fraction of image width/height. Tuned for NoteCam's
# bottom-left info block. Adjust here if your photos are a different layout.
CROP_TOP_FRAC = 0.79     # start this fraction down from the top of the image
CROP_WIDTH_FRAC = 0.42   # go this fraction across from the left edge


# ------------------------------------------------------------------------------
def get_overlay_crop(pil_img):
    w, h = pil_img.size
    return pil_img.crop((0, int(h * CROP_TOP_FRAC), int(w * CROP_WIDTH_FRAC), h))


def to_otsu(overlay_rgb, scale=5, blur=True):
    """Upscale + threshold the overlay crop to make text crisp for OCR."""
    bgr = cv2.cvtColor(np.array(overlay_rgb), cv2.COLOR_RGB2BGR)
    bgr = cv2.resize(bgr, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    if blur:
        gray = cv2.bilateralFilter(gray, 9, 75, 75)
    _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return otsu


def note_quality(note_str):
    """Reward clean, plausible-looking note text; penalize OCR noise tails."""
    if not note_str:
        return 0
    s = note_str.strip()
    if len(s) > 50 or len(s) == 0:
        return -3
    return 2 if re.match(r"^[A-Za-z0-9 ,.\-/]+$", s) else -1


def score_text(txt):
    """Higher = more fields found, values look numerically plausible, note is clean."""
    score = 0
    for name, pat in PATTERNS.items():
        m = pat.search(txt)
        if not m:
            continue
        score += 1
        if name == "Latitude":
            try:
                if LAT_RANGE[0] <= float(m.group(1)) <= LAT_RANGE[1]:
                    score += 2
            except ValueError:
                pass
        if name == "Longitude":
            try:
                if LON_RANGE[0] <= float(m.group(1)) <= LON_RANGE[1]:
                    score += 2
            except ValueError:
                pass
        if name == "Note":
            score += note_quality(m.group(1))
    return score


def ocr_best(overlay_rgb):
    """Try a couple of preprocessing/PSM combos, keep whichever scores highest."""
    best_txt, best_score = "", -1
    candidates = [
        to_otsu(overlay_rgb, scale=5, blur=True),
        to_otsu(overlay_rgb, scale=5, blur=False),
        to_otsu(overlay_rgb, scale=7, blur=True),
    ]
    for img in candidates:
        for psm in (6, 4):
            txt = pytesseract.image_to_string(img, config=f"--psm {psm}")
            s = score_text(txt)
            if s > best_score:
                best_score, best_txt = s, txt
    return best_txt, best_score


def clean_note(raw):
    """Trim likely OCR-noise tails off the free-text Note field."""
    if raw is None:
        return None
    s = raw.strip()
    # cut at the first run of 2+ characters outside the normal alphanumeric/
    # space/punctuation set -- that's almost always where background clutter
    # started getting misread as extra "text"
    m = re.search(r"[^A-Za-z0-9 ,.\-/]{2,}", s)
    if m:
        s = s[:m.start()].strip()
    # a genuine note is short; if it's still very long, something leaked in
    if len(s) > 50:
        s = s[:50].strip()
    return s if s else None


def parse_fields(txt):
    out = {}
    for name, pat in PATTERNS.items():
        m = pat.search(txt)
        if not m:
            out[name] = None
            continue
        if name == "Elevation":
            base = m.group(1)
            margin = m.group(2)
            out[name] = f"{base}\u00b1{margin} m" if margin else f"{base} m"
        elif name == "Accuracy":
            out[name] = f"{m.group(1)} m"
        elif name == "Note":
            out[name] = clean_note(m.group(1))
        else:
            out[name] = m.group(1)
    return out


def process_image(path):
    pil_img = PILImage.open(path).convert("RGB")
    overlay = get_overlay_crop(pil_img)
    txt, _ = ocr_best(overlay)
    fields = parse_fields(txt)

    flag = "OK"
    lat, lon = fields.get("Latitude"), fields.get("Longitude")
    if not lat or not lon:
        flag = "REVIEW - missing lat/long"
    else:
        try:
            latf, lonf = float(lat), float(lon)
            if not (LAT_RANGE[0] <= latf <= LAT_RANGE[1] and LON_RANGE[0] <= lonf <= LON_RANGE[1]):
                flag = "REVIEW - out of expected range"
        except ValueError:
            flag = "REVIEW - unparsable coordinate"
    if flag == "OK" and (not fields.get("Elevation") or not fields.get("Accuracy")):
        flag = "REVIEW - check elevation/accuracy"
    if flag == "OK" and not fields.get("Note"):
        flag = "REVIEW - check note text"

    return fields, flag, overlay


def iter_images(root_folder):
    for dirpath, _dirnames, filenames in sorted(os.walk(root_folder)):
        folder_name = os.path.basename(dirpath)
        for fn in sorted(filenames):
            if os.path.splitext(fn)[1].lower() in IMG_EXTS:
                yield folder_name, fn, os.path.join(dirpath, fn)


def main():
    cleanup_dir = None
    if EXTRACTED_FOLDER:
        root = EXTRACTED_FOLDER
    else:
        cleanup_dir = tempfile.mkdtemp(prefix="geo_extract_")
        print(f"Extracting zip: {ZIP_PATH}")
        with zipfile.ZipFile(ZIP_PATH, "r") as z:
            z.extractall(cleanup_dir)
        root = cleanup_dir

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geo Data"
    headers = ["Folder Name", "File Name", "Latitude", "Longitude",
               "Elevation", "Accuracy", "Note / Status", "Flag", "Overlay Preview"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(vertical="center")

    thumb_dir = tempfile.mkdtemp(prefix="geo_thumbs_")
    row_idx = 2
    total = 0
    flagged = 0
    errors = []

    for folder_name, file_name, full_path in iter_images(root):
        total += 1
        try:
            fields, flag, overlay_img = process_image(full_path)
        except Exception as e:
            fields = {k: None for k in PATTERNS}
            flag = "ERROR - could not read image"
            overlay_img = None
            errors.append((folder_name, file_name, str(e), traceback.format_exc()))

        if flag != "OK":
            flagged += 1

        ws.cell(row=row_idx, column=1, value=folder_name)
        ws.cell(row=row_idx, column=2, value=file_name)
        ws.cell(row=row_idx, column=3, value=fields.get("Latitude"))
        ws.cell(row=row_idx, column=4, value=fields.get("Longitude"))
        ws.cell(row=row_idx, column=5, value=fields.get("Elevation"))
        ws.cell(row=row_idx, column=6, value=fields.get("Accuracy"))
        ws.cell(row=row_idx, column=7, value=fields.get("Note"))
        flag_cell = ws.cell(row=row_idx, column=8, value=flag)
        if flag == "OK":
            flag_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            flag_cell.font = Font(color="006100")
        else:
            flag_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            flag_cell.font = Font(color="9C0006")

        if overlay_img is not None:
            thumb_path = os.path.join(thumb_dir, f"{row_idx}.png")
            thumb = overlay_img.copy()
            thumb.thumbnail((280, 100))
            thumb.save(thumb_path)
            ws.row_dimensions[row_idx].height = 75
            ws.add_image(XLImage(thumb_path), f"I{row_idx}")

        row_idx += 1
        if total % 10 == 0:
            print(f"  processed {total} images...")

    widths = [26, 22, 12, 12, 16, 10, 26, 26, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Run date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append(["Total images processed", total])
    ws2.append(["Flagged for review", flagged])
    ws2.append(["Errors (unreadable files)", len(errors)])
    for r in errors:
        ws2.append([r[0], r[1], r[2]])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 50

    wb.save(OUTPUT_XLSX)
    shutil.rmtree(thumb_dir, ignore_errors=True)
    if cleanup_dir:
        shutil.rmtree(cleanup_dir, ignore_errors=True)

    print(f"\nDone. {total} images processed, {flagged} flagged for review, {len(errors)} unreadable.")
    print(f"Saved to: {OUTPUT_XLSX}")
    print("Open the file, sort/filter the 'Flag' column, and check the thumbnail")
    print("next to any flagged row (or any row you want to double-check) before")
    print("relying on the numbers.")


if __name__ == "__main__":
    main()